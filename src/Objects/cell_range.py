import re
import pandas as pd
from Objects.cell_reference import CellReference
from openpyxl.utils import column_index_from_string, get_column_letter

class CellRange:
    pattern = r"([A-Z]+\d+):([A-Z]+\d+)"  # Class attribute for the regex pattern

    @staticmethod
    def is_valid_range(range_str):
        return bool(re.match(CellRange.pattern, range_str)) or isinstance(range_str, CellRange)

    def __init__(self, start_ref, end_ref=None):
        if end_ref is None:
            self.start_cell, self.end_cell = self.parse_range(start_ref)
        else:
            self.start_cell = CellReference(start_ref)
            self.end_cell = CellReference(end_ref)

    def parse_range(self, range_str):
        match = re.match(CellRange.pattern, range_str)
        if not match:
            raise ValueError(f"Invalid range format: {range_str}")
        start_ref, end_ref = match.groups()
        return CellReference(start_ref), CellReference(end_ref)
    
    def get_rows_in_range(self):
        start_row = self.start_cell.row_number
        end_row = self.end_cell.row_number
        return list(range(start_row, end_row + 1))

    def get_columns_in_range(self, as_numbers=False):
        start_col = column_index_from_string(self.start_cell.column_letter)
        end_col = column_index_from_string(self.end_cell.column_letter)
        if as_numbers:
            return list(range(start_col, end_col + 1))
        else:
            return [get_column_letter(col) for col in range(start_col, end_col + 1)]
        
    def get_cells_in_range(self, as_dataframe=False):
        start_col = column_index_from_string(self.start_cell.column_letter)
        end_col = column_index_from_string(self.end_cell.column_letter)
        start_row = self.start_cell.row_number
        end_row = self.end_cell.row_number

        cells = [[f"{get_column_letter(col)}{row}" for col in range(start_col, end_col + 1)] for row in range(start_row, end_row + 1)]

        if as_dataframe:
            return pd.DataFrame(cells)
        return cells


    def to_dict(self):
        # Create dictionary representation of the range
        return {
            "cell_range": {
                "start": str(self.start_cell),
                "end": str(self.end_cell)
            }
        }

    def __str__(self):
        return f"{self.start_cell}:{self.end_cell}"
    

if __name__ == "__main__":
    valid_ranges = ['A1:B2', 'C3:D4', 'E5:F6', 'G7:H8']
    invalid_ranges = ['A1B2', 'A1:', ':B2', '1A:B2', 'A1:2B']

    for valid_range in valid_ranges:
        try:
            cell_range = CellRange(valid_range)
            print(f"Successfully created cell range: {cell_range}")
        except ValueError as e:
            print(f"Error creating cell range: {e}")

    for invalid_range in invalid_ranges:
        try:
            cell_range = CellRange(invalid_range)
            print(f"Successfully created cell range: {cell_range}")
        except ValueError as e:
            print(f"Error creating cell range: {e}")