import re
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter
from Models.reference import Reference

class Range:
    pattern = r"([A-Z]+\d+):([A-Z]+\d+)"  # Class attribute for the regex pattern

    @staticmethod
    def from_dict(range_dict):
        """Reconstruct the cell range from its dictionary representation."""
        start_ref = range_dict['components']['start']
        if 'reference' in start_ref:
            start_ref = str(Reference.from_dict(start_ref))
        end_ref = range_dict['components']['end']
        if 'reference' in end_ref:
            end_ref = str(Reference.from_dict(end_ref))
        return Range(f"{start_ref}:{end_ref}")

    @staticmethod
    def is_valid_range(range_str):
        """Check if the string represents a valid cell range."""
        return bool(re.match(Range.pattern, range_str)) or isinstance(range_str, Range)

    def __init__(self, start_ref, end_ref=None):
        """Initialize the range either from a range string or from two references."""
        if end_ref is None:
            self.start_cell, self.end_cell = self.parse_range(start_ref)
        else:
            self.start_cell = Reference(start_ref)
            self.end_cell = Reference(end_ref)
        self.rows = self.get_rows_in_range()
        self.columns = self.get_columns_in_range(as_numbers=True)
        self.cell_matrix = self.get_cells_in_range()  # Get a list of lists of cell references
        self.row_index = 0
        self.col_index = 0

    def parse_range(self, range_str):
        """Parse a range string into start and end CellReferences."""
        match = re.match(Range.pattern, range_str)
        if not match:
            raise ValueError(f"Invalid range format: {range_str}")
        start_ref, end_ref = match.groups()
        return Reference(start_ref), Reference(end_ref)
    
    def get_rows_in_range(self):
        """Return a list of rows covered by the range."""
        return list(range(self.start_cell.row_number, self.end_cell.row_number + 1))

    def get_columns_in_range(self, as_numbers=False):
        """Return a list of columns covered by the range, as numbers or letters."""
        start_col = column_index_from_string(self.start_cell.column_letter)
        end_col = column_index_from_string(self.end_cell.column_letter)
        if as_numbers:
            return list(range(start_col, end_col + 1))
        else:
            return [get_column_letter(col) for col in range(start_col, end_col + 1)]
        
    def get_cells_in_range(self, as_dataframe=False):
        """Return all cells in the range, optionally as a pandas DataFrame."""
        rows = range(self.start_cell.row_number, self.end_cell.row_number + 1)
        cols = range(column_index_from_string(self.start_cell.column_letter), column_index_from_string(self.end_cell.column_letter) + 1)
        cells = [[f"{get_column_letter(col)}{row}" for col in cols] for row in rows]

        if as_dataframe:
            return pd.DataFrame(cells, index=[f"Row {row}" for row in rows], columns=[get_column_letter(col) for col in cols])
        return cells

    def to_dict(self):
        """Create a dictionary representation of the cell range."""
        return {
            "range": str(self),
            "components": {
                "start": str(self.start_cell),
                "end": str(self.end_cell)
            }
        }

    def __iter__(self):
        """Reset the iterator to the start of the range."""
        self.row_index = 0
        self.col_index = 0
        return self

    def __next__(self):
        if self.row_index >= len(self.rows):
            raise StopIteration

        current_cell = self.cell_matrix[self.row_index][self.col_index]
        self.col_index += 1

        if self.col_index >= len(self.columns):
            self.col_index = 0
            self.row_index += 1

        return current_cell

    def __contains__(self, item):
        """Check if a cell like 'B1' is in the range of 'A1:B2' for example."""
        if isinstance(item, str):
            try:
                item_ref = Reference(item)
                return (column_index_from_string(self.start_cell.column_letter) <= column_index_from_string(item_ref.column_letter) <= column_index_from_string(self.end_cell.column_letter)) and (self.start_cell.row_number <= item_ref.row_number <= self.end_cell.row_number)
            except ValueError:
                return False
        return False

    def __str__(self):
        """String representation of the cell range."""
        return f"{self.start_cell}:{self.end_cell}"

if __name__ == "__main__":
    valid_ranges = ['A1:B2', 'C3:D4', 'E5:F6', 'G7:H8']
    invalid_ranges = ['A1B2', 'A1:', ':B2', '1A:B2', 'A1:2B']

    for valid_range in valid_ranges:
        try:
            cell_range = Range(valid_range)
            print(f"Successfully created cell range: {cell_range}")
        except ValueError as e:
            print(f"Error creating cell range: {e}")

    for invalid_range in invalid_ranges:
        try:
            cell_range = Range(invalid_range)
            print(f"Successfully created cell range: {cell_range}")
        except ValueError as e:
            print(f"Error creating cell range: {e}")
